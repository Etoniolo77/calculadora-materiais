#!/usr/bin/env python3
"""
Reimporta a aba "Lista Consolidada" do Excel mestre para o Supabase,
preservando a coluna **Aplicação** (por material) e parseando quantidades
em texto ("4,5MTS", "2,4KG (15MTS)").

Modelo: a Aplicação é por MATERIAL, não por estrutura. Por isso:
- `estruturas` passa a ter UMA linha por código (tipo_poste = 'ALL').
- `estrutura_materiais` ganha a coluna `aplicacao` (texto da Aplicação).
- A seleção por tipo de poste é feita em runtime por `core.aplicacao.aplicacao_matches`.

Faz backup das tabelas antes de reescrever. Idempotente.
"""
from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import psycopg2
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "core"))
from aplicacao import parse_qty_text  # noqa: E402

EXCEL_PATH = PROJECT_ROOT / "docs" / "Referencia_Tecnica" / "ESTRUTURAS PARA CALCULADORA MATERIAS.xlsx"
SHEET = "Lista Consolidada"
BACKUP_DIR = PROJECT_ROOT / "data" / "_backup_mestre"
MATERIAL_RE = re.compile(r"^\d{8}$")


def _read_excel() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[SHEET]
    header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: header.index(name) for name in header}
    col_est = idx.get("Estrutura")
    col_cod = idx.get("Código Hana", idx.get("Codigo Hana"))
    col_desc = idx.get("Descrição", idx.get("Descricao"))
    col_qty = idx.get("Quantidade")
    col_apl = idx.get("Aplicação", idx.get("Aplicacao"))
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        est = str(row[col_est] or "").strip().upper()
        cod = str(row[col_cod] or "").strip().split(".")[0]
        desc = str(row[col_desc] or "").strip()
        qty = parse_qty_text(row[col_qty])
        apl = str(row[col_apl] or "").strip()
        if not est or not MATERIAL_RE.match(cod) or not desc or qty <= 0:
            continue
        out.append({"estrutura": est, "codigo": cod, "desc": desc, "qty": qty, "aplicacao": apl})
    wb.close()
    return out


def _backup(cur) -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    for table in ("estruturas", "estrutura_materiais"):
        cur.execute(f"SELECT * FROM {table}")
        cols = [d[0] for d in cur.description]
        data = [dict(zip(cols, r)) for r in cur.fetchall()]
        path = BACKUP_DIR / f"{table}_backup_{stamp}.json"
        path.write_text(json.dumps(data, default=str, ensure_ascii=False), encoding="utf-8")
        print(f"[BACKUP] {table}: {len(data)} linhas -> {path.name}")


def main() -> int:
    load_dotenv(PROJECT_ROOT / "backend" / ".env")
    rows = _read_excel()
    estruturas = sorted({r["estrutura"] for r in rows})
    print(f"[EXCEL] {len(rows)} linhas validas, {len(estruturas)} estruturas")

    conn = psycopg2.connect(os.environ["SUPABASE_DB_URL"])
    try:
        cur = conn.cursor()
        _backup(cur)
        cur.execute(
            "ALTER TABLE estrutura_materiais ADD COLUMN IF NOT EXISTS aplicacao TEXT"
        )
        # Rebuild transacional.
        cur.execute("DELETE FROM estrutura_materiais")
        cur.execute("DELETE FROM estruturas")
        code_to_id: dict[str, int] = {}
        for codigo in estruturas:
            cur.execute(
                "INSERT INTO estruturas (codigo, tipo_poste) VALUES (%s, 'ALL') RETURNING id",
                (codigo,),
            )
            code_to_id[codigo] = cur.fetchone()[0]
        args = [
            (code_to_id[r["estrutura"]], r["codigo"], r["desc"], r["qty"], r["aplicacao"])
            for r in rows
        ]
        cur.executemany(
            """
            INSERT INTO estrutura_materiais
                (estrutura_id, material_codigo, material_descricao, quantidade, aplicacao)
            VALUES (%s, %s, %s, %s, %s)
            """,
            args,
        )
        conn.commit()
        cur.execute("SELECT count(*) FROM estruturas")
        n_est = cur.fetchone()[0]
        cur.execute("SELECT count(*) FROM estrutura_materiais")
        n_mat = cur.fetchone()[0]
        print(f"[OK] estruturas={n_est}, estrutura_materiais={n_mat}")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
