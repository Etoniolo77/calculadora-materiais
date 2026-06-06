"""
rebuild_database.py
===================
Reconstrói o materials.db e unified_db.json a partir de TODAS as fontes:

Fontes lidas (em ordem de prioridade para conflitos):
  0. EXPORT_DB_*.xlsx      → baseline completo (6650 materiais, 251 estruturas)
  1. Matriz original.xlsx  → sobreposição/atualização de estruturas × postes
  2. Estruturas *.pdf      → confirma/complementa a matriz
  3. Cintas.pdf            → lookup de diâmetros de cintas por poste
  4. Materiais Trafo.pdf   → kits de hardware de transformadores
  5. Codigos de Materiais Novos.xlsx → catálogo SAP complementar

Saída:
  data/unified_db.json  → fonte JSON (manter sempre atualizada)
  data/materials.db     → banco SQLite pronto para uso pelo app

Executar da raiz do projeto:
  python scripts/rebuild_database.py
"""

import json
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber

# ── Caminhos ──────────────────────────────────────────────────────────────────
ROOT        = Path(__file__).parent.parent
DATA_DIR    = ROOT / "data"
BASES_DIR   = DATA_DIR / "Bases_Dados"
REF_DIR     = ROOT / "docs" / "Referencia_Tecnica"
OUT_JSON    = DATA_DIR / "unified_db.json"
OUT_DB      = DATA_DIR / "materials.db"

# ── Helpers ───────────────────────────────────────────────────────────────────

def clean_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def clean_num(v, default=1.0) -> float:
    try:
        return float(str(v).replace(",", ".").split()[0])
    except Exception:
        return default

def pole_type_from_col(col_header: str) -> str:
    """
    Extrai o tipo de poste de cabeçalhos como 'N1 - D11/300' → 'DT'
    ou 'B2 - C12/600' → 'CIRCULAR'.
    """
    h = col_header.upper()
    if re.search(r'\bD\d{2}/\d{3,4}\b', h):
        return "DT"
    if re.search(r'\bC\d{1,2}/\d{3,4}\b', h):
        return "CIRCULAR"
    return "ALL"

def struct_code_from_col(col_header: str) -> str:
    """
    Extrai o código da estrutura de 'N1 - D11/300' → 'N1'.
    Remove espaços internos: 'CE 2' → 'CE2'.
    """
    parts = col_header.split(" - ")
    code = parts[0].strip().replace(" ", "")
    return code

# ── 0. EXPORT_DB baseline (EXPORT_DB_*.xlsx) ─────────────────────────────────

def load_export_db(bases_dir: Path) -> tuple:
    """
    Lê o export do banco original (EXPORT_DB_*.xlsx).
    Retorna (sap_catalog, export_structs) onde:
      sap_catalog    = {codigo: descricao}  (até 6650 materiais)
      export_structs = {codigo: [{'sap':..,'desc':..,'qty':..,'type':..}, ...]}
    """
    sap_catalog = {}
    export_structs = {}

    candidates = sorted(bases_dir.glob("EXPORT_DB_*.xlsx"))
    if not candidates:
        print("  [AVISO] Nenhum EXPORT_DB_*.xlsx encontrado")
        return sap_catalog, export_structs

    fpath = candidates[-1]  # Usar o mais recente
    print(f"  [{fpath.name}] carregando...")

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

        # MATERIAIS sheet
        if "MATERIAIS" in wb.sheetnames:
            ws = wb["MATERIAIS"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                code = clean_str(row[0]).split(".")[0]
                desc = clean_str(row[1]) if len(row) > 1 else ""
                if code and desc and not code.startswith("9"):
                    sap_catalog[code] = desc
            print(f"  [{fpath.name}] {len(sap_catalog)} materiais carregados de MATERIAIS")

        # ESTRUTURA_MATS sheet
        if "ESTRUTURA_MATS" in wb.sheetnames:
            ws = wb["ESTRUTURA_MATS"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                struct = clean_str(row[0])
                ptype  = clean_str(row[1]) or "ALL"
                sap    = clean_str(row[2]).split(".")[0]
                desc   = clean_str(row[3]) if len(row) > 3 else ""
                qty    = clean_num(row[4]) if len(row) > 4 else 1.0
                if not struct or not sap:
                    continue
                desc = desc or sap_catalog.get(sap, f"MAT {sap}")
                if struct not in export_structs:
                    export_structs[struct] = []
                entry = {"sap": sap, "desc": desc, "qty": qty, "type": ptype}
                if entry not in export_structs[struct]:
                    export_structs[struct].append(entry)
            n_items = sum(len(v) for v in export_structs.values())
            print(f"  [{fpath.name}] {len(export_structs)} estruturas, {n_items} itens carregados de ESTRUTURA_MATS")

        wb.close()
    except Exception as e:
        print(f"  [ERRO] {fpath.name}: {e}")

    return sap_catalog, export_structs


# ── 1. Catálogo SAP (Codigos de Materiais Novos.xlsx) ────────────────────────

def load_sap_catalog(bases_dir: Path) -> dict:
    """Retorna {codigo: descricao} a partir das planilhas de códigos SAP."""
    catalog = {}

    for fname in ["Codigos de Materiais Novos.xlsx", "Codigos de Materiais.xlsx"]:
        fpath = bases_dir / fname
        if not fpath.exists():
            print(f"  [AVISO] {fname} não encontrado")
            continue

        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb["CODIGOS"]
        header = [clean_str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        # Identificar colunas
        col_new  = next((i for i, h in enumerate(header) if "Material Novo" in h or "Material" in h), 0)
        col_desc = next((i for i, h in enumerate(header) if "Texto" in h or "Descri" in h), 2)

        for row in ws.iter_rows(min_row=2, values_only=True):
            code = clean_str(row[col_new]).split(".")[0]
            desc = clean_str(row[col_desc])
            if code and desc and code.isdigit():
                catalog[code] = desc

        wb.close()
        print(f"  [{fname}] {len(catalog)} materiais carregados")

    return catalog


# ── 2. Estruturas via Matriz original.xlsx ────────────────────────────────────

def load_structures_from_matrix(bases_dir: Path, sap_catalog: dict) -> dict:
    """
    Lê a aba MATRIZ do arquivo e retorna:
      { 'N1': [{'sap':..., 'desc':..., 'qty':..., 'type':'CIRCULAR'}, ...], ... }

    A matriz tem:
      Col 0: CODIGO (SAP)
      Col 1: MATERIAIS (descrição)
      Col 2+: uma por estrutura+poste_tipo (ex: 'N1 - C12/600'), valor = quantidade
    """
    structures = {}
    fpath = bases_dir / "Matriz original.xlsx"
    if not fpath.exists():
        print("  [AVISO] Matriz original.xlsx não encontrada")
        return structures

    wb  = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws  = wb["MATRIZ"]

    # Ler cabeçalho completo
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers    = [clean_str(h) for h in header_row]

    # Mapear colunas de estrutura (índice → (struct_code, pole_type))
    col_map = {}
    for i, h in enumerate(headers[2:], start=2):
        if not h or " - " not in h:
            continue
        struct = struct_code_from_col(h)
        ptype  = pole_type_from_col(h)
        if struct:
            col_map[i] = (struct, ptype)
            if struct not in structures:
                structures[struct] = []

    print(f"  [MATRIZ] {len(col_map)} colunas de estrutura × poste mapeadas")
    print(f"  [MATRIZ] Estruturas únicas: {sorted(set(v[0] for v in col_map.values()))}")

    # Ler linhas de materiais
    rows_processed = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        sap_raw = clean_str(row[0]).split(".")[0]
        if not sap_raw or not re.match(r"^\d{5,10}$", sap_raw):
            continue

        desc_raw = clean_str(row[1]) if len(row) > 1 else ""
        desc = desc_raw or sap_catalog.get(sap_raw, f"MAT {sap_raw}")

        # Para cada coluna de estrutura, verificar quantidade
        for col_idx, (struct, ptype) in col_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == 0 or val == "" or val is False:
                continue
            try:
                qty = float(str(val).replace(",", "."))
            except Exception:
                continue
            if qty <= 0:
                continue

            # Regras de negócio: Sela → DT, Cinta → CIRCULAR
            effective_type = ptype
            desc_upper = desc.upper()
            if "SELA" in desc_upper:
                effective_type = "DT"
            elif "CINTA" in desc_upper or "BRAÇADEIRA" in desc_upper:
                effective_type = "CIRCULAR"

            # Evitar duplicata exata
            entry = {"sap": sap_raw, "desc": desc, "qty": qty, "type": effective_type}
            if entry not in structures[struct]:
                structures[struct].append(entry)

        rows_processed += 1

    wb.close()
    print(f"  [MATRIZ] {rows_processed} linhas de material processadas")
    print(f"  [MATRIZ] {sum(len(v) for v in structures.values())} entradas totais")
    return structures


# ── 3. Estruturas via PDFs (Estruturas *.pdf) ─────────────────────────────────

def load_structures_from_pdfs(ref_dir: Path, sap_catalog: dict) -> dict:
    """
    Lê os PDFs de Estruturas (Estruturas N.pdf, Estruturas B.pdf, etc.)
    via extração de tabelas. Retorna o mesmo formato de structures.
    """
    structures = {}
    pdf_files = sorted(ref_dir.glob("Estruturas *.pdf"))
    print(f"  [PDFs] {len(pdf_files)} PDFs de estrutura encontrados")

    for pdf_path in pdf_files:
        print(f"    Processando {pdf_path.name}...")
        try:
            with pdfplumber.open(pdf_path) as pdf:
                current_structure = None
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        for row in table:
                            if not any(row):
                                continue
                            row_vals = [clean_str(x) for x in row]

                            # Detectar header → skip
                            if "ESTRUTURA" in row_vals[0].upper() or (len(row_vals) > 4 and "MATERIAL" in row_vals[4].upper()):
                                continue

                            # Atualizar código de estrutura (forward fill)
                            cand = row_vals[0]
                            if cand and len(cand) < 7 and not cand.lower().startswith("ver"):
                                m = re.match(r"^([A-Z0-9\-]+)", cand.upper())
                                if m:
                                    current_structure = m.group(1)
                                    if current_structure not in structures:
                                        structures[current_structure] = []

                            if not current_structure:
                                continue

                            # Buscar código SAP (8 dígitos) em cols 1 ou 2
                            code, qty, desc, idx_sap = None, 1.0, "", -1
                            for i in [1, 2]:
                                if i >= len(row_vals):
                                    continue
                                val = row_vals[i]
                                if re.match(r"^\d{8}$", val):
                                    code = val
                                    idx_sap = i
                                    break
                                if "ver. POSTE" in val or "ver. CABO" in val:
                                    code = "VERIFICAR-" + val.replace("ver. ", "")
                                    idx_sap = i
                                    break

                            if not code:
                                continue

                            if idx_sap + 1 < len(row_vals):
                                qty = clean_num(row_vals[idx_sap + 1])
                            if idx_sap + 2 < len(row_vals):
                                desc = row_vals[idx_sap + 2] or sap_catalog.get(code, "")

                            # Circular/DT das últimas 2 colunas
                            col_circ = row_vals[-2] if len(row_vals) >= 2 else ""
                            col_dt   = row_vals[-1] if len(row_vals) >= 1 else ""

                            has_circ = bool(re.match(r"^[Xx\d]", col_circ)) and col_circ.strip() not in ("", "0")
                            has_dt   = bool(re.match(r"^[Xx\d]", col_dt))   and col_dt.strip()   not in ("", "0")

                            qty_dt = qty
                            if re.match(r"^\d+[xX]", col_dt):
                                try:
                                    qty_dt = float(col_dt.split("x")[0].split("X")[0])
                                    has_dt = True
                                except Exception:
                                    pass

                            desc_upper = desc.upper()
                            if "SELA" in desc_upper:
                                has_circ, has_dt = False, True
                            elif "CINTA" in desc_upper or "BRAÇADEIRA" in desc_upper:
                                has_circ, has_dt = True, False

                            def _add(struct, sap, d, q, t):
                                entry = {"sap": sap, "desc": d, "qty": q, "type": t}
                                if entry not in structures[struct]:
                                    structures[struct].append(entry)
                                if sap not in sap_catalog and not sap.startswith("VERIF"):
                                    sap_catalog[sap] = d

                            if has_circ and has_dt:
                                if qty_dt != qty:
                                    _add(current_structure, code, desc, qty,    "CIRCULAR")
                                    _add(current_structure, code, desc, qty_dt, "DT")
                                else:
                                    _add(current_structure, code, desc, qty, "ALL")
                            elif has_circ:
                                _add(current_structure, code, desc, qty, "CIRCULAR")
                            elif has_dt:
                                _add(current_structure, code, desc, qty_dt, "DT")

        except Exception as e:
            print(f"    [ERRO] {pdf_path.name}: {e}")

    print(f"  [PDFs] {len(structures)} estruturas extraídas dos PDFs")
    return structures


# ── 4. Cintas lookup (Cintas.pdf) ─────────────────────────────────────────────

def load_cinta_lookup(ref_dir: Path) -> dict:
    """
    Extrai o mapeamento poste → diâmetro de cinta do PDF Cintas.pdf.
    Retorna { 'C12/600': {'CINTA 1': 240, ...}, ... }
    """
    lookup = {}
    fpath = ref_dir / "Cintas.pdf"
    if not fpath.exists():
        print("  [AVISO] Cintas.pdf não encontrado")
        return lookup

    try:
        with pdfplumber.open(fpath) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue
                    # Header: primeira linha com tipos de poste
                    header = [clean_str(c) for c in (table[0] or [])]
                    for row in table[1:]:
                        if not row or not row[0]:
                            continue
                        cat_raw = clean_str(row[0]).upper()  # ex: "CINTA 1", "ESTAI 1"
                        cat = None
                        if "CINTA" in cat_raw:
                            cat = "CINTA 1"
                        elif "ESTAI" in cat_raw:
                            cat = "ESTAI 1"
                        elif "NIVEL" in cat_raw:
                            cat = "NIVEL 1"
                        elif "RECK" in cat_raw or "REK" in cat_raw:
                            cat = "RECK 1"
                        elif "SECUNDARIA" in cat_raw:
                            cat = "SECUNDARIA"
                        elif "LUMINARIA" in cat_raw:
                            cat = "LUMINARIA"
                        if not cat:
                            continue

                        for i, h in enumerate(header[1:], start=1):
                            if not h or i >= len(row):
                                continue
                            # Normalizar chave de poste
                            pole_key = h.strip().upper().replace(" ", "")
                            val = clean_str(row[i])
                            try:
                                diameter = int(float(val.replace(",", ".")))
                                if pole_key not in lookup:
                                    lookup[pole_key] = {}
                                lookup[pole_key][cat] = diameter
                            except Exception:
                                pass
    except Exception as e:
        print(f"  [ERRO] Cintas.pdf: {e}")

    print(f"  [Cintas.pdf] {len(lookup)} postes mapeados")
    return lookup


# ── 5. Kits de Hardware de Trafo (Materiais Trafo.pdf) ───────────────────────

def load_trafo_kits(ref_dir: Path, sap_catalog: dict) -> dict:
    """
    Extrai kits de hardware de transformadores do PDF.
    Retorna { 'TRAFO_MONO': [...], 'TRAFO_TRI_45': [...], ... }
    """
    kits = {
        # Kits fixos sempre presentes
        "ESTAI_CC_14M": [
            {"sap": "30056363", "desc": "HASTE ANCOR AC 1020 3200DAN 16MM 1,6M", "qty": 1},
            {"sap": "30054507", "desc": "CORDOALHA ACO CARB 9,5MM 7F CL.B MR/SM", "qty": 14},
        ],
        "PARA_RAIO_CONJUNTO": [
            {"sap": "30053319", "desc": "COBERTURA PROT PARA-RAIO 13,8KV", "qty": 1},
        ],
    }

    fpath = ref_dir / "Materiais Trafo.pdf"
    if not fpath.exists():
        print("  [AVISO] Materiais Trafo.pdf não encontrado")
        return kits

    try:
        with pdfplumber.open(fpath) as pdf:
            current_kit = None
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not any(row):
                            continue
                        row_vals = [clean_str(x) for x in row]
                        text_join = " ".join(row_vals).upper()

                        # Detectar seção do kit
                        if "MONO" in text_join and ("KIT" in text_join or "TRAFO" in text_join):
                            current_kit = "TRAFO_MONO"
                            if current_kit not in kits:
                                kits[current_kit] = []
                            continue
                        if "TRIF" in text_join and ("KIT" in text_join or "TRAFO" in text_join):
                            current_kit = "TRAFO_TRI_45"
                            if current_kit not in kits:
                                kits[current_kit] = []
                            continue

                        if not current_kit:
                            continue

                        # Extrair SAP
                        for i, val in enumerate(row_vals):
                            if re.match(r"^\d{8}$", val):
                                sap = val
                                qty = clean_num(row_vals[i + 1]) if i + 1 < len(row_vals) else 1.0
                                desc = row_vals[i + 2] if i + 2 < len(row_vals) else sap_catalog.get(sap, "")
                                desc = desc or sap_catalog.get(sap, f"MAT {sap}")
                                entry = {"sap": sap, "desc": desc, "qty": qty}
                                if entry not in kits[current_kit]:
                                    kits[current_kit].append(entry)
                                break
    except Exception as e:
        print(f"  [ERRO] Materiais Trafo.pdf: {e}")

    print(f"  [Materiais Trafo.pdf] {len(kits)} kits de hardware")
    return kits


# ── 6. Merge de estruturas (EXPORT_DB + MATRIZ + PDF) ────────────────────────

def merge_structures(matrix_structs: dict, pdf_structs: dict, export_structs: dict) -> dict:
    """
    Combina estruturas em 3 camadas de prioridade:
      1. EXPORT_DB baseline (todas as 251 estruturas originais)
      2. Matriz original.xlsx sobreposição por SAP (dados mais recentes)
      3. PDFs complemento (SAPs ausentes nas fontes acima)
    """
    # Camada 0: começar com export_structs como base
    merged = {k: list(v) for k, v in export_structs.items()}

    def _overlay(source: dict, overwrite: bool):
        """Adiciona/atualiza merged com entradas de source."""
        for struct, items in source.items():
            if struct not in merged:
                merged[struct] = list(items)
                continue
            existing_saps = {e["sap"] for e in merged[struct]}
            for item in items:
                if item["sap"] not in existing_saps:
                    merged[struct].append(item)
                    existing_saps.add(item["sap"])
                elif overwrite:
                    # Atualizar qty/desc para entradas existentes
                    for e in merged[struct]:
                        if e["sap"] == item["sap"] and e.get("type") == item.get("type"):
                            e["qty"]  = item["qty"]
                            e["desc"] = item["desc"] or e["desc"]
                            break

    # Camada 1: MATRIZ sobreposição (atualiza qtds)
    _overlay(matrix_structs, overwrite=True)

    # Camada 2: PDFs complemento (apenas adiciona ausentes)
    _overlay(pdf_structs, overwrite=False)

    return merged


# ── 7. Construir unified_db.json ─────────────────────────────────────────────

def build_unified_db(structures: dict, sap_catalog: dict, cinta_lookup: dict, hw_kits: dict) -> dict:
    return {
        "metadata": {
            "version": "3.0",
            "description": "Unified DB rebuilt from all sources (Matrix + PDFs + Excel)",
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "sources": [
                "Matriz original.xlsx (MATRIZ)",
                "Estruturas *.pdf",
                "Cintas.pdf",
                "Materiais Trafo.pdf",
                "Codigos de Materiais Novos.xlsx",
            ],
        },
        "sap_library": sap_catalog,
        "structures": structures,
        "cinta_lookup": cinta_lookup,
        "hardware_kits": hw_kits,
    }


# ── 8. Construir materials.db ─────────────────────────────────────────────────

def build_sqlite(db_path: Path, unified_db: dict):
    if db_path.exists():
        db_path.unlink()
        print(f"  Banco anterior removido: {db_path}")

    conn = sqlite3.connect(db_path)
    cur  = conn.cursor()

    # Schema
    cur.executescript("""
        CREATE TABLE materiais (
            codigo   TEXT PRIMARY KEY,
            descricao TEXT NOT NULL
        );
        CREATE VIRTUAL TABLE materiais_fts
            USING fts5(codigo, descricao, content='materiais', content_rowid='rowid');
        CREATE TRIGGER materiais_ai AFTER INSERT ON materiais BEGIN
            INSERT INTO materiais_fts(rowid, codigo, descricao) VALUES (new.rowid, new.codigo, new.descricao);
        END;
        CREATE TRIGGER materiais_ad AFTER DELETE ON materiais BEGIN
            INSERT INTO materiais_fts(materiais_fts, rowid, codigo, descricao)
                VALUES('delete', old.rowid, old.codigo, old.descricao);
        END;
        CREATE TRIGGER materiais_au AFTER UPDATE ON materiais BEGIN
            INSERT INTO materiais_fts(materiais_fts, rowid, codigo, descricao)
                VALUES('delete', old.rowid, old.codigo, old.descricao);
            INSERT INTO materiais_fts(rowid, codigo, descricao) VALUES (new.rowid, new.codigo, new.descricao);
        END;
        CREATE TABLE estruturas (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo      TEXT NOT NULL,
            tipo_poste  TEXT DEFAULT 'ALL'
        );
        CREATE INDEX idx_est_cod ON estruturas(codigo);
        CREATE TABLE estrutura_materiais (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            estrutura_id       INTEGER NOT NULL,
            material_codigo    TEXT NOT NULL,
            material_descricao TEXT,
            quantidade         REAL NOT NULL,
            FOREIGN KEY (estrutura_id) REFERENCES estruturas(id)
        );
        CREATE INDEX idx_em_estid ON estrutura_materiais(estrutura_id);
    """)

    # Materiais SAP
    n_mat = 0
    for code, desc in unified_db["sap_library"].items():
        if code and desc and not str(code).startswith("9"):
            cur.execute("INSERT OR REPLACE INTO materiais (codigo, descricao) VALUES (?,?)", (code, desc))
            n_mat += 1
    print(f"  {n_mat} materiais inseridos")

    # Estruturas e seus materiais
    n_est = 0
    n_em  = 0
    for struct_code, items in unified_db["structures"].items():
        # Agrupar por tipo de poste
        by_type: dict = {}
        for item in items:
            t = item.get("type", "ALL")
            by_type.setdefault(t, []).append(item)

        for ptype, mats in by_type.items():
            cur.execute("INSERT INTO estruturas (codigo, tipo_poste) VALUES (?,?)", (struct_code, ptype))
            est_id = cur.lastrowid
            n_est += 1
            for m in mats:
                cur.execute(
                    "INSERT INTO estrutura_materiais (estrutura_id, material_codigo, material_descricao, quantidade) VALUES (?,?,?,?)",
                    (est_id, m["sap"], m.get("desc", ""), m.get("qty", 1)),
                )
                n_em += 1

    conn.commit()

    # Validar FTS
    test = conn.execute("SELECT codigo, descricao FROM materiais_fts WHERE materiais_fts MATCH '\"POSTE\"*' LIMIT 2").fetchall()
    conn.close()

    print(f"  {n_est} entradas de estrutura inseridas ({len(unified_db['structures'])} estruturas únicas)")
    print(f"  {n_em} materiais de estrutura inseridos")
    print(f"  FTS test OK: {len(test)} resultado(s) para 'POSTE'")


# ── Relatório de validação ────────────────────────────────────────────────────

def print_report(unified_db: dict):
    structs = unified_db["structures"]
    sap_lib = unified_db["sap_library"]

    print("\n" + "=" * 60)
    print("RELATORIO DE VALIDACAO")
    print("=" * 60)
    print(f"  Materiais SAP      : {len(sap_lib)}")
    print(f"  Estruturas unicas  : {len(structs)}")
    print(f"  Total itens mat*est: {sum(len(v) for v in structs.values())}")
    print()

    # Estruturas sem material
    vazias = [k for k, v in structs.items() if not v]
    if vazias:
        print(f"  [!] Estruturas SEM materiais ({len(vazias)}): {sorted(vazias)}")
    else:
        print("  [OK] Todas as estruturas tem pelo menos 1 material")

    # Estruturas com VERIFICAR
    com_verif = [k for k, v in structs.items() if any("VERIF" in str(m.get("sap","")) for m in v)]
    if com_verif:
        print(f"  [!] Estruturas com VERIFICAR ({len(com_verif)}): {sorted(com_verif)}")

    # Listar todas
    print(f"\n  Estruturas ({len(structs)}):")
    for code in sorted(structs.keys()):
        n = len(structs[code])
        types = sorted({m.get("type","?") for m in structs[code]})
        print(f"    {code:15} {n:3} materiais  [{', '.join(types)}]")

    print("=" * 60)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Forçar encoding UTF-8 no stdout para suportar caracteres especiais
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    print("\n" + "=" * 60)
    print("REBUILD DATABASE - PRJ-13-Calculadora")
    print(f"Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 0. EXPORT_DB baseline
    print("\n[0/6] Carregando EXPORT_DB baseline...")
    sap_catalog, export_structs = load_export_db(BASES_DIR)

    # 1. Catálogo SAP complementar
    print("\n[1/6] Carregando catalogo SAP complementar...")
    sap_extra = load_sap_catalog(BASES_DIR)
    # Merge: export tem prioridade, sap_extra preenche gaps
    for code, desc in sap_extra.items():
        if code not in sap_catalog:
            sap_catalog[code] = desc
    print(f"  Catalogo final: {len(sap_catalog)} materiais")

    # 2. Estruturas via Matriz
    print("\n[2/6] Lendo Matriz original.xlsx...")
    matrix_structs = load_structures_from_matrix(BASES_DIR, sap_catalog)

    # 3. Estruturas via PDFs
    print("\n[3/6] Lendo PDFs de estrutura...")
    pdf_structs = load_structures_from_pdfs(REF_DIR, sap_catalog)

    # 4. Merge (export + matrix + pdfs)
    print("\n[4/6] Mesclando fontes (EXPORT + MATRIZ + PDFs)...")
    structures = merge_structures(matrix_structs, pdf_structs, export_structs)
    print(f"  {len(structures)} estruturas unicas apos merge")

    # 5. Cintas e Kits
    print("\n[5/6] Lendo Cintas.pdf e Materiais Trafo.pdf...")
    cinta_lookup = load_cinta_lookup(REF_DIR)
    hw_kits = load_trafo_kits(REF_DIR, sap_catalog)

    # 6. Salvar
    print("\n[6/6] Salvando unified_db.json e materials.db...")
    unified_db = build_unified_db(structures, sap_catalog, cinta_lookup, hw_kits)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(unified_db, f, indent=2, ensure_ascii=False)
    print(f"  unified_db.json salvo: {OUT_JSON}")

    build_sqlite(OUT_DB, unified_db)
    print(f"  materials.db  salvo: {OUT_DB}")

    print_report(unified_db)
    print(f"\nRebuild concluido: {datetime.now().strftime('%H:%M:%S')}")


if __name__ == "__main__":
    main()
